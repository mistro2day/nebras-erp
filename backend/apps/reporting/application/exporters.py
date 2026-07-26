# -*- coding: utf-8 -*-
"""
مُصدِّرات التقارير إلى PDF وExcel بدعم عربي كامل (RTL + تشكيل الحروف).

- PDF: reportlab + Amiri (OFL مُضمّن) + arabic_reshaper + python-bidi.
  يعمل على ويندوز (التطوير) ولينكس (الإنتاج) دون مكتبات نظام.
- Excel: openpyxl بورقة RTL وترويسة منسّقة.

الاستخدام عبر ExportService في services.py.
"""
import io
import os

import arabic_reshaper
from bidi.algorithm import get_display

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

_FONT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts')
_FONT_REGISTERED = False

# ألوان هوية نبراس (متوافقة مع رموز التصميم في الواجهة)
_BRAND = colors.HexColor('#0f766e')       # teal الأساسي
_BRAND_DARK = colors.HexColor('#134e4a')
_HEADER_BG = colors.HexColor('#0f766e')
_ROW_ALT = colors.HexColor('#f1f5f9')
_BORDER = colors.HexColor('#cbd5e1')
_TEXT = colors.HexColor('#0f172a')
_MUTED = colors.HexColor('#64748b')


def _register_fonts():
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return
    pdfmetrics.registerFont(TTFont('Amiri', os.path.join(_FONT_DIR, 'Amiri-Regular.ttf')))
    pdfmetrics.registerFont(TTFont('Amiri-Bold', os.path.join(_FONT_DIR, 'Amiri-Bold.ttf')))
    pdfmetrics.registerFontFamily('Amiri', normal='Amiri', bold='Amiri-Bold')
    _FONT_REGISTERED = True


def ar(text):
    """يشكّل النص العربي ويطبّق الاتجاه ثنائي الجهة للعرض الصحيح في reportlab."""
    if text is None:
        return ''
    s = str(text)
    # الأرقام والنصوص اللاتينية تُترك كما هي؛ التشكيل يعالج العربية داخلها
    try:
        return get_display(arabic_reshaper.reshape(s))
    except Exception:
        return s


class PdfExporter:
    """يبني ملف PDF أنيقاً لتقرير: ترويسة + جدول بيانات + تذييل."""

    @classmethod
    def build(cls, report_name, rows, tenant_name='', subtitle='', generated_at=''):
        _register_fonts()
        buf = io.BytesIO()

        # اتجاه الصفحة: أفقي إن كثرت الأعمدة
        cols = list(rows[0].keys()) if rows else []
        pagesize = landscape(A4) if len(cols) > 5 else A4

        doc = SimpleDocTemplate(
            buf, pagesize=pagesize,
            rightMargin=14 * mm, leftMargin=14 * mm,
            topMargin=16 * mm, bottomMargin=16 * mm,
            title=report_name,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ArTitle', parent=styles['Title'], fontName='Amiri-Bold',
            fontSize=18, textColor=_BRAND_DARK, alignment=TA_RIGHT, leading=24)
        sub_style = ParagraphStyle(
            'ArSub', fontName='Amiri', fontSize=10.5, textColor=_MUTED,
            alignment=TA_RIGHT, leading=16)
        cell_style = ParagraphStyle(
            'ArCell', fontName='Amiri', fontSize=9.5, textColor=_TEXT,
            alignment=TA_RIGHT, leading=13)
        head_style = ParagraphStyle(
            'ArHead', fontName='Amiri-Bold', fontSize=10, textColor=colors.white,
            alignment=TA_CENTER, leading=14)

        elements = []

        # ترويسة: اسم المؤسسة + عنوان التقرير + الوصف + تاريخ التوليد
        if tenant_name:
            elements.append(Paragraph(ar(tenant_name), sub_style))
        elements.append(Paragraph(ar(report_name), title_style))
        if subtitle:
            elements.append(Paragraph(ar(subtitle), sub_style))
        if generated_at:
            # التاريخ لاتيني، فنبنيه معكوساً ليظهر صحيحاً بعد التشكيل
            elements.append(Paragraph(ar('تاريخ التوليد: ') + generated_at, sub_style))
        elements.append(Spacer(1, 8 * mm))

        if not rows:
            elements.append(Paragraph(ar('لا توجد بيانات لعرضها في هذا التقرير.'), cell_style))
            doc.build(elements)
            return buf.getvalue()

        # الجدول: نعكس ترتيب الأعمدة لتبدأ من اليمين (RTL)
        rcols = list(reversed(cols))
        header = [Paragraph(ar(c), head_style) for c in rcols]
        table_data = [header]
        for row in rows:
            table_data.append([
                Paragraph(ar(row.get(c, '')), cell_style) for c in rcols
            ])

        avail = doc.width
        col_w = avail / len(rcols)
        table = Table(table_data, colWidths=[col_w] * len(rcols), repeatRows=1)
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), _HEADER_BG),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, _BORDER),
            ('LINEBELOW', (0, 0), (-1, 0), 1, _BRAND_DARK),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                style.append(('BACKGROUND', (0, i), (-1, i), _ROW_ALT))
        table.setStyle(TableStyle(style))
        elements.append(table)
        elements.append(Spacer(1, 6 * mm))
        elements.append(Paragraph(
            ar(f'إجمالي السجلات: {len(rows)}'),
            ParagraphStyle('ArTot', fontName='Amiri-Bold', fontSize=10,
                           textColor=_BRAND_DARK, alignment=TA_RIGHT)))

        doc.build(elements, onFirstPage=cls._footer, onLaterPages=cls._footer)
        return buf.getvalue()

    @staticmethod
    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Amiri', 8)
        canvas.setFillColor(_MUTED)
        page = ar(f'صفحة {doc.page}')
        canvas.drawString(14 * mm, 8 * mm, 'Nebras ERP')
        canvas.drawRightString(doc.pagesize[0] - 14 * mm, 8 * mm, page)
        canvas.restoreState()


class ExcelExporter:
    """يبني ملف Excel بورقة RTL وترويسة منسّقة."""

    @classmethod
    def build(cls, report_name, rows, tenant_name='', generated_at=''):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'التقرير'
        ws.sheet_view.rightToLeft = True

        cols = list(rows[0].keys()) if rows else []
        ncol = max(len(cols), 1)

        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill('solid', fgColor='0F766E')
        alt_fill = PatternFill('solid', fgColor='F1F5F9')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center')

        # صف العنوان (مدموج)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        c = ws.cell(row=1, column=1, value=report_name)
        c.font = Font(name='Arial', size=15, bold=True, color='134E4A')
        c.alignment = center
        # سطر المعلومات
        info = tenant_name
        if generated_at:
            info = (info + '  ·  ' if info else '') + f'تاريخ التوليد: {generated_at}'
        if info:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
            ic = ws.cell(row=2, column=1, value=info)
            ic.font = Font(name='Arial', size=10, color='64748B')
            ic.alignment = center

        start = 4
        if not rows:
            ws.cell(row=start, column=1, value='لا توجد بيانات لعرضها.')
            return cls._save(wb)

        # الترويسة
        for j, col in enumerate(cols, start=1):
            cell = ws.cell(row=start, column=j, value=col)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # الصفوف
        for i, row in enumerate(rows, start=1):
            r = start + i
            for j, col in enumerate(cols, start=1):
                val = row.get(col, '')
                cell = ws.cell(row=r, column=j, value=val)
                cell.font = Font(name='Arial', size=10, color='0F172A')
                cell.alignment = right if isinstance(val, str) else center
                cell.border = border
                if i % 2 == 0:
                    cell.fill = alt_fill

        # عرض الأعمدة تقريبي
        for j, col in enumerate(cols, start=1):
            width = max(len(str(col)), *(len(str(row.get(col, ''))) for row in rows)) + 4
            ws.column_dimensions[get_column_letter(j)].width = min(max(width, 12), 40)

        return cls._save(wb)

    @staticmethod
    def _save(wb):
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
